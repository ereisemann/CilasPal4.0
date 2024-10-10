import sys
sys.path.append('/Users/eveeisemann/Documents/GitHub/CilasPal4.0/CilasPal')
import PdfReaderObj as p
import CilasPalSetup as setup
import CilasPalDebugger as debug
import pypdf
import xlsxwriter
import openpyxl
import numpy as np
import math
import os
import re   ### ~ere
from colorama import Fore

def debug(file, defined_classes_headers, standard_classes_headers):
    pdf = p.PdfReaderObject(file)

    spreadsheet_path = pdf.init_excel()
    print(Fore.YELLOW + f"spreadsheet path is {spreadsheet_path}")
    print(Fore.YELLOW + f"Size classes are hard coded as follows: Standard classes = {standard_classes_headers}, Customer defined classes = {defined_classes_headers}")

    row = 2
    # Now parse info from pdf. I'll keep this as straight forward as possible
    for i in range(0, pdf.num_pages, 2):
        contents1 = pdf.read_content(i)
        contents2 = pdf.read_content(i+1)
        sample_name = contents1[contents1.index("Sample ref.") + 14:contents1.index("Sample Name")]
        median = contents1[contents1.index('Diameter at 50%') + 18:contents1.index('µmDiameter at 90%')-1]
        mean = contents1[contents1.index('Mean diameter') + 16:contents1.index('FraunhoferDensity')-4]

        # Check index alignment on first sample
        if i >= 1000:
            print(Fore.YELLOW + "Whoa thats a lot of samples! ...Quitting Debug Mode...")
        else:
            print(Fore.YELLOW + f"...Testing Index Alignment on sample {i} with ID {sample_name}...\n")
            print(f"name:{sample_name}mean: {mean}\nmedian: {median}")

            if " " in sample_name or " " in mean or " " in median: # If there are spaces before/after
                print(Fore.RED + "Misalignment in indexing size metrics. Check that pdf version == or that the sample name does not contain spaces. ...Continuing with incorrect indexing to shwo error effects...")
            else:
                print(Fore.GREEN + "Correct indexing size metrics ... checking size classes..." + Fore.WHITE)

                contents1 = contents1[620:]    

                defined_class_distrib = []
                print(Fore.YELLOW + "...solving Indexing problem with start=anchor+6, end=anchor+11 assuming Cilas outputs values as length 5 FPNs")
                for classs in defined_classes_headers:
                    s = contents1.index(classs)
                    try:
                        val = float(contents1[s+6:s+11])     # Catches if output contains characters --> can not convert a misallignment to float
                        defined_class_distrib.append(val)
                    except:
                        print(Fore.RED + f"Misalignment in indexing size classes. Can not cast {contents1[s+6:s+11]} to float ...continuing with bad solution...")
                        defined_class_distrib.append(contents1[s+6:s+11])   ### contents 1 indexing can stay the same b/c only two rows
                print(Fore.YELLOW + f"Customer defined ouput: {defined_class_distrib}")

                #contents2 = contents2[620:]  ### previous trimming
                search_pattern_1 = '\nx\nQ3\nq3'    ### start ### improved string trimming based on pattern ~ere
                search_pattern_2 = 'diameter'   ### end ###
                if search_pattern_1 and search_pattern_2 in contents2:

                    start_index = contents2.index(search_pattern_1) ### first occurrence pattern 1
                    end_index = contents2.rfind(search_pattern_2) ### last occurrence of pattern 2

                    trimmed_txt = updated_text[start_index:end_index] ### Trim everything outside those
                    trimmed_text = trimmed_txt.replace('Q3', ' ').replace('q3', ' ')
                else:
                    trimmed_text = contents2   ### Result if no patterns found, probably unnecessary

                cleaned_text = re.findall(r'[0-9]+\.[0-9]+(?:\.[0-9]+)?', trimmed_text)  ### isolating numbers, decimals, spaces
                #cleaned_string = ' '.join(cleaned_text)   ### reassembling ### REMOVE

                def split_long_strings(data):    ### Function to split numbers that are mashed together. ~ere
                    result = []
                    for item in data:
                        ### If the string is longer than 5 characters, split
                        if len(item) > 5 and item != '100.00':  ### where item length exceeds 5 and are not 100.00
                            ### Check if the string contains a decimal after the first character
                            if item[1] == '.':
                                split_point = 4
                            else:
                                split_point = 3
                            ### Split the string at the calculated point and append both parts to result
                            result.append(item[:split_point])
                            result.append(item[split_point:])
                        else:  ### when the string is 5 characters or less or equal to 100.00, keep unsplit
                            result.append(item)
                    ### Return the list of strings, excluding spaces
                    return [s for s in result if s.strip()]

                split_text = split_long_strings(cleaned_text)   ### reassign contents2 to the final version of the list of strings
                #contents2 = ' '.join(split_text)### REMOVE make it back into one giant string with spaces so the next loop works, probably change this eventually to index a vector rather than a mega string?

                standard_class_distrib = []

                print(sample_name)

                for classs in standard_classes_headers:
                    #s = contents2.index(classs)
                    s = split_text.index(classs)   ### indexing from the string list instead, now s, s+1, s+2 are the three vals.
                    ###
                    #print(str(classs) + " = SIZE CLASS " + str(s) + " = STR INDEX")   ###
                    #print(split_text[s:s+3])
                    print(split_text[s] + ' = CLASS ' + split_text[s+1] + " = CUMULATIVE " + split_text[s+2] + " = NONCUMULATIVE ")
                    ###

                    try:
                        #val = float(contents2[s+12:s+17])   ### ~ere
                        val = float(split_text[s+2])  ### from cleaned up list of strings
                        standard_class_distrib.append(val)
                    except:
                        print(Fore.RED + f"Misalignment in indexing size classes. Can not cast {contents2[s+12:s+17]} to float ...continuing with bad solution...")  ### ~ere
                        #standard_class_distrib.append(contents2[s+12:s+17]) ### ~ere
                        standard_class_distrib.append(split_text[s+2])  ### ~ere
                print(Fore.YELLOW + f"Standard defined output: {standard_class_distrib}")

                workbook = openpyxl.load_workbook(spreadsheet_path)
                defined_class_distrib_sheet = workbook[workbook.sheetnames[1]]
                standard_class_distrib_sheet = workbook[workbook.sheetnames[0]]

            standard_class_distrib_sheet.cell(row=row, column=1, value=sample_name)
            standard_class_distrib_sheet.cell(row=row, column=2, value=mean)
            standard_class_distrib_sheet.cell(row=row, column=3, value=median)
            for val, j in zip(standard_class_distrib, range(0, len(standard_class_distrib))):
                standard_class_distrib_sheet.cell(row=row, column=j+4, value=val)
            workbook.save(spreadsheet_path)
            defined_class_distrib_sheet.cell(row=row, column=1, value=sample_name)
            for val, j in zip(defined_class_distrib, range(0, len(defined_class_distrib))):
                defined_class_distrib_sheet.cell(row=row, column=j+2, value=val)
            workbook.save(spreadsheet_path)
            row+=1