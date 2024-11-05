#import sys
#sys.path.append('/Users/eveeisemann/Documents/GitHub/CilasPal4.0/CilasPal')   ## If script can't find PdfReaderObj while workingin code, map your directory here
#sys.path.append(r'C:\Users\eveve\Documents\Git\CilasPal4.0\CilasPal')   ## If script can't find PdfReaderObj map your directory here
import PdfReaderObj as p
import pypdf
import xlsxwriter
import openpyxl
import numpy as np
import math
import os
import re   ### ~ere
from colorama import Fore
import warnings

# Global vars
defined_classes_headers = ['0.04', '3.90', '62.00', '88.00', '125.0', '177.0', '250.0', '350.0', '500.0', '710.0', '1000.0',
                                        '1410.0', '2000.0']

standard_classes_headers = ['0.04', '0.07', '0.10', '0.20', '0.30', '0.40', '0.50', '0.60', '0.70', '0.80', '0.90', '1.00', '1.10',
                                        '1.20', '1.30', '1.40', '1.60', '1.80', '2.00', '2.20', '2.40', '2.60', '3.00', '4.00', '5.00', '6.00', '6.50', '7.00',
                                        '7.50', '8.00', '8.50', '9.00', '10.00', '11.00', '12.00', '13.00', '14.00', '15.00', '16.00', '17.00', '18.00',
                                        '19.00', '20.00', '22.00', '25.00', '28.00', '32.00', '36.00', '38.00', '40.00', '45.00', '50.00', '53.00', '56.00',
                                        '63.00', '71.00', '75.00', '80.00', '85.00', '90.00', '95.00', '100.0', '106.0', '112.0', '125.0', '130.0',
                                        '140.0', '145.0', '150.0', '160.0', '170.0', '180.0', '190.0', '200.0', '212.0', '242.0', '250.0',
                                        '300.0', '400.0', '500.0', '600.0', '700.0', '800.0', '900.0', '1000.0', '1100.0', '1200.0', '1300.0',
                                        '1400.0', '1500.0', '1600.0', '1700.0', '1800.0', '1900.0', '2000.0', '2100.0', '2200.0', '2300.0',
                                        '2400.0', '2500.0']

file = None

debug = input("Run in debug mode? (Y/N)")
if debug == "Y" or debug == "y":
    from CilasPalSetup import build_env
    from CilasPalDebugger import debug
    file = "tests/testfile.pdf"
    print(Fore.YELLOW + "...debugging with verbose output...")
    print(Fore.YELLOW + "...rebuilding CilasPal env...")
    build_env()
    print(Fore.YELLOW + "...running test file...")
    debug(file, defined_classes_headers, standard_classes_headers)
else:
    print("Welcome, please paste the path to the data file without quotes (.pdf only!)")   ### it does not seem to accept a file path with quotes ~ere
    file = input("Full file path is: ")
    file = file.strip('"').strip("'")   ### allows input to include quotes around file path.

    pdf = p.PdfReaderObject(file)      ## PDF defined here based on PDF reader object (see PdfReaderObj.py)

    spreadsheet_path = pdf.init_excel()
    print(spreadsheet_path)

    row = 2
    # Now parse info from pdf. I'll keep this as straight forward as possible
    for i in range(0, pdf.num_pages, 2):
        contents1 = pdf.read_content(i)   ## Contents of page 1 of sample i (user defined size classes)
        contents2 = pdf.read_content(i+1) ## Contents of page 2 of sample i (all size classes)
        sample_name = contents1[contents1.index("Sample ref.") + 14:contents1.index("Sample Name")]
        median = contents1[contents1.index('Diameter at 50%') + 18:contents1.index('µmDiameter at 90%')-1]
        mean = contents1[contents1.index('Mean diameter') + 16:contents1.index('FraunhoferDensity')-4]

        print(sample_name)  ## keep in some verbose for regular script
    #
    #     # Check index alignment on first sample
    #     if i == 0:
    #         print(Fore.YELLOW + "...Testing Index Alignment on sample 1...\n" + Fore.WHITE)
    #         print(f"name:{sample_name}mean: {mean}\nmedian: {median}")
    #
    #         if " " in sample_name or " " in mean or " " in median: # If there are spaces before/after
    #             #raise Exception("Misalignment in indexing size metrics. Check that pdf version == or that the sample name does not contain spaces.")   ### Now that indexing is based on a cleaned version of all the values, not sure spaces in sample names are a problem
    #             warnings.warn("Misalignment in indexing size metrics. Check that pdf version == or that the sample name does not contain spaces.")   ### Now that indexing is based on a cleaned version of all the values, not sure spaces in sample names are a problem
    #         else:
    #             print(Fore.GREEN + "Correct indexing size metrics ... checking size classes..." + Fore.WHITE)
    #
    #             contents1 = contents1[620:]
    #
    #             defined_class_distrib = []
    #             for classs in defined_classes_headers:
    #                 s = contents1.index(classs)
    #                 try:
    #                     val = float(contents1[s+6:s+11])     # Catches if output contains characters --> can not convert a misalignment to float
    #                     defined_class_distrib.append(val)
    #                 except:
    #                     print(Fore.RED + f"Misalignment in indexing size classes. Can not cast {contents1[s+6:s+11]} to float" + Fore.WHITE)
    #                     raise Exception("Misalignment in indexing size classes. Can not cast to float. see above")
    #             print(Fore.GREEN + "Correct indexing for customer defined size classes ... checking standard classes..." + Fore.WHITE)
    #
    #
    # ### major update below ~ere
    #             search_pattern_1 = '\nx\nQ3\nq3'    ### start ### improved string trimming based on pattern ~ere
    #             search_pattern_2 = 'diameter'   ### end ###
    #             if search_pattern_1 and search_pattern_2 in contents2:
    #                 start_index = contents2.index(search_pattern_1) ### first occurrence pattern 1
    #                 end_index = contents2.rfind(search_pattern_2) ### last occurrence of pattern 2
    #                 trimmed_txt = contents2[start_index:end_index] ### Trim everything outside those
    #                 trimmed_text = trimmed_txt.replace('Q3', ' ').replace('q3', ' ')
    #             else:
    #                 trimmed_text = contents2   ### Result if no patterns found, probably unnecessary
    #
    #             cleaned_text = re.findall(r'[0-9]+\.[0-9]+(?:\.[0-9]+)?', trimmed_text)  ### isolating numbers, decimals, spaces
    #
    #             def split_long_strings(data):    ### Function to split numbers that are mashed together. ~ere
    #                 result = []
    #                 for item in data:
    #                     ### If the string is longer than 5 characters, split
    #                     if len(item) > 6 and item != '100.00':  ### where item length exceeds 6 and are not 100.00
    #                         ### Check if the string contains a decimal after the first character
    #                         if item[1] == '.':
    #                             split_point = 4
    #                         else:
    #                             split_point = 3
    #                         ### Split the string at the calculated point and append both parts to result
    #                         result.append(item[:split_point])
    #                         result.append(item[split_point:])
    #                     else:  ### when the string is 5 characters or less or equal to 100.00, keep unsplit
    #                         result.append(item)
    #                 ### Return the list of strings, excluding spaces
    #                 return [s for s in result if s.strip()]
    #
    #             split_text = split_long_strings(cleaned_text)   ### reassign contents2 to the final version of the list of strings
    # ### major update above ~ere
    #
    #             standard_class_distrib = []
    #             for j in range(1, len(standard_classes_headers)):  ### sometimes class header shows up in the data before appropriate location, causing problems with this indexing method
    #                 try:
    #                     val = float(split_text[j * 3 - 1])
    #                     standard_class_distrib.append(val)
    #                 except:
    #                     print(Fore.RED + f"Misalignment in indexing size classes. Can not cast {split_text[j * 3 - 1]} to float" + Fore.WHITE)  ### ~ere
    #                     raise Exception("Misalignment in indexing size classes. Can not cast to float, see above")
    #             print(Fore.GREEN + "Correct indexing for standard size classes ... proceeding to parse full data..." + Fore.WHITE)


        # Same exact thing as above, still catches misalignment errors but has no verbose output   ### Commenting this out b/c unneeded ~ere
        contents1 = pdf.read_content(i)
        contents2 = pdf.read_content(i+1)
        if " " in sample_name or " " in mean or " " in median:
            # raise Exception("Misalignment in indexing size metrics. Check that pdf version == or that the sample name does not contain spaces.")   ### Now that indexing is based on a cleaned version of all the values, not sure spaces in sample names are a problem
            warnings.warn("Misalignment in indexing size metrics. Check that pdf version == or that the sample name does not contain spaces.")  ### Now that indexing is based on a cleaned version of all the values, not sure spaces in sample names are a problem
        else:
            contents1 = contents1[620:]
            defined_class_distrib = []
            for classs in defined_classes_headers:
                s = contents1.index(classs)
                try:
                    val = float(contents1[s+6:s+11])
                    defined_class_distrib.append(val)
                except:
                    print(Fore.RED + f"Misalignment in indexing size classes. Can not cast {contents1[s+6:s+11]} to float" + Fore.WHITE)
                    raise Exception("Misalignment in indexing size classes. Can not cast to float. see above")

 ### major update below ~ere

            search_pattern_1 = '\nx\nQ3\nq3'  ### start ### improved string trimming based on pattern ~ere
            search_pattern_2 = 'diameter'  ### end ###
            if search_pattern_1 and search_pattern_2 in contents2:
                start_index = contents2.index(search_pattern_1)  ### first occurrence pattern 1
                end_index = contents2.rfind(search_pattern_2)  ### last occurrence of pattern 2
                trimmed_txt = contents2[start_index:end_index]  ### Trim everything outside those
                trimmed_text = trimmed_txt.replace('Q3', ' ').replace('q3', ' ')
            else:
                trimmed_text = contents2  ### Result if no patterns found, probably unnecessary

            cleaned_text = re.findall(r'[0-9]+\.[0-9]+(?:\.[0-9]+)?',
                                      trimmed_text)  ### isolating numbers, decimals, spaces

            def split_long_strings(data):  ### Function to split numbers that are mashed together. ~ere
                result = []
                for item in data:
                    ### If the string is longer than 5 characters, split
                    if len(item) > 6 and item != '100.00':  ### where item length exceeds 6 and are not 100.00
                        ### Check if the string contains a decimal after the first character
                        if item[1] == '.':
                            split_point = 4
                        else:
                            split_point = 3
                        ### Split the string at the calculated point and append both parts to result
                        result.append(item[:split_point])
                        result.append(item[split_point:])
                    else:  ### when the string is 5 characters or less or equal to 100.00, keep unsplit
                        result.append(item)
                ### Return the list of strings, excluding spaces
                return [s for s in result if s.strip()]

            split_text = split_long_strings(cleaned_text)  ### reassign contents2 to the final version of the list of strings
 ### major update above ~ere

            standard_class_distrib = []
            #for classs in standard_classes_headers:
            for j in range(1,len(standard_classes_headers)):
                try:
                    val = float(split_text[j * 3 - 1])
                    standard_class_distrib.append(val)
                except:
                    print(Fore.RED + f"Misalignment in indexing size classes. Can not cast {split_text[j * 3 - 1]} to float" + Fore.WHITE)  ### ~ere
                    raise Exception("Misalignment in indexing size classes. Can not cast to float, see above")


            workbook = openpyxl.load_workbook(spreadsheet_path)
            defined_class_distrib_sheet = workbook[workbook.sheetnames[1]]
            standard_class_distrib_sheet = workbook[workbook.sheetnames[0]]


        standard_class_distrib_sheet.cell(row=row, column=1, value=sample_name)
        standard_class_distrib_sheet.cell(row=row, column=2, value=mean)
        standard_class_distrib_sheet.cell(row=row, column=3, value=median)
        for val, j in zip(standard_class_distrib, range(0, len(standard_class_distrib))):
            standard_class_distrib_sheet.cell(row=row, column=j+4, value=val)
        workbook.save(spreadsheet_path)
        defined_class_distrib_sheet.cell(row=row, column=1, value=sample_name)
        for val, j in zip(defined_class_distrib, range(0, len(defined_class_distrib))):
            defined_class_distrib_sheet.cell(row=row, column=j+2, value=val)
        workbook.save(spreadsheet_path)
        row+=1
