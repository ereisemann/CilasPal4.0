'''
Standards naming convention is firstSecondThird for class names and first_second_third for function names. Variables all follow the python standard of first_second_etc except in some special cases. To test PdfReader class run the file to run unit tests in main()
'''

from pypdf import PdfReader
from pathlib import Path
import os
import xlsxwriter
import openpyxl

class PdfReaderObject:
    def __init__(self, fileName):
        # Corner Cases
        if not fileName.endswith(".pdf"):
            raise Exception("Can not construct a PdfReader object from a non-.pdf file")
        if not os.path.exists(fileName):
            raise Exception("Invalid path argument. Ensure you are passing the full file path")
            
        self.path_string = fileName
        self.path_object = Path(fileName)
        self.reader = PdfReader(fileName)
        self.num_pages = len(self.reader.pages)
        self.spreadsheet_path_ref = None
        
        # Need even pages
        if self.num_pages%2 != 0:
            raise Exception(f"pdf passed has an odd length: {self.num_pages} not divisble by 2. Check that no samples are missing")
        
    def read_content(self, page_num):
        # Corner cases
        if page_num < 0 or page_num > self.num_pages:
            raise Exception(f"Can not read content from page {page_num} when pdf has length {self.num_pages}")
            
        page = self.reader.pages[page_num]
        return page.extract_text()
    
    def pages(self):
        return num_pages
    
    def init_excel(self, overwrite=False):
        spreadsheet_path = str(self.path_object.parent) + '/Your_Data.xlsx'    ###
        self.spreadsheet_path_ref = spreadsheet_path
        if os.path.exists(spreadsheet_path):
            if overwrite:
                print(f"Overwriting existing file: '{spreadsheet_path}'")
                os.remove(spreadsheet_path)
            else:
                raise FileExistsError(f"the file '{spreadsheet_path}' already exists. Rename or delete and re-run CilasPal.")   ### throw file exists error to clarify for users
        #elif (not os.path.exists(spreadsheet_path)):
        workbook = xlsxwriter.Workbook(spreadsheet_path)
        standard_classes = workbook.add_worksheet("Standard_Classes")
        defined_classes = workbook.add_worksheet("Customer_Defined_Classes")
        standard_classes.write(0, 0, 'x')
        workbook.close()
            
            # handle headers first
        standard_classes_headers = ["ID", "Mean", "Median", 0.04, 0.07, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0, 1.1,
           1.2, 1.3, 1.4, 1.6, 1.8, 2.0, 2.2, 2.4, 2.6, 3.0, 4.0, 5.0, 6.0, 6.5, 7.0,
           7.5, 8.0, 8.5, 9.0, 10.0, 11.0, 12.0, 13.0, 14.0, 15.0, 16.0, 17.0, 18.0,
           19.0, 20.0, 22.0, 25.0, 28.0, 32.0, 36.0, 38.0, 40.0, 45.0, 50.0, 53.0, 56.0,
           63.0, 71.0, 75.0, 80.0, 85.0, 90.0, 95.0, 100.0, 106.0, 112.0, 125.0, 130.0,
           140.0, 145.0, 150.0, 160.0, 170.0, 180.0, 190.0, 200.0, 212.0, 242.0, 250.0,
           300.0, 400.0, 500.0, 600.0, 700.0, 800.0, 900.0, 1000.0, 1100.0, 1200.0, 1300.0,
           1400.0, 1500.0, 1600.0, 1700.0, 1800.0, 1900.0, 2000.0, 2100.0, 2200.0, 2300.0,
           2400.0, 2500.0]
        defined_classes_headers = ["ID", 0.04, 3.90, 62.00, 88.00, 125.00, 177.0, 2500.0, 350.0, 500.0, 710.0, 1000.0, 1410.0, 2000.0]
            
        workbook = openpyxl.load_workbook(spreadsheet_path)

        # Load each page
        standard_classes_page = workbook[workbook.sheetnames[0]]
        defined_classes_page = workbook[workbook.sheetnames[1]]
            # Add the headers
        i = 1
        for defined_header in defined_classes_headers:
            defined_classes_page.cell(row=1, column=i, value=defined_header)
            i+=1
             
        i=1
        for standard_header in standard_classes_headers:
            standard_classes_page.cell(row=1, column=i, value=standard_header)
            i+=1
                
        workbook.save(spreadsheet_path)
        return spreadsheet_path
            
    def main():
        sample_path = '/Users/coltenrodriguez/Desktop/fun_cilas_adventure/test_data/test_data.pdf'
        pdf = PdfReaderObject(sample_path)
        
                
                
            
            
            
        